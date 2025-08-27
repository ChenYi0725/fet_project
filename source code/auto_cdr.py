from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QHeaderView, QMessageBox
from auto_cdr_ui import Ui_MainWindow 
import excel_info 
import openpyxl
import shutil
from openpyxl.styles import Font

class MainWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.ui.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.ui.output_button.clicked.connect(self.on_output_clicked)
        self.init_table()

    def init_table(self):
        for row in range(self.ui.tableWidget.rowCount()):
            for col in range(self.ui.tableWidget.columnCount()):
                self.ui.tableWidget.setItem(row, col, QtWidgets.QTableWidgetItem(""))

    def get_selected_band(self):
        selected_band = []
        if self.ui.radioButton_700.isChecked():
            selected_band.append("700")
        if self.ui.radioButton_700.isChecked():
            selected_band.append("700 5M")
        if self.ui.radioButton_1800.isChecked():
            selected_band.append("1800")
        if self.ui.radioButton_2100.isChecked():
            selected_band.append("2100")
        if self.ui.radioButton_2600.isChecked():
            selected_band.append("2600")
        if self.ui.radioButton_2600_tdd.isChecked():
            selected_band.append("2600 TDD")
        return selected_band

    def on_output_clicked(self):
        node_name = self.ui.node_name_input.text()
        tac = self.ui.tac_input.text()
        longitude = self.ui.longitude_input.text()
        latitude = self.ui.node_name_input_5.text()
        bands = self.get_selected_band()
        input_info = excel_info.ExcelInfo(
           node_name=node_name,
           tac=tac,
           longitude=longitude,
           latitude=latitude,
           bands=bands,
           table_widget=self.ui.tableWidget,
        )
     
        if input_info.is_input_correct():              
            self.save_to_excel(export_map=input_info.export_map)
            
            self.clear_interface()
        else:
            self.show_message("輸入錯誤", "請檢查您的輸入是否正確。", QMessageBox.Warning)

    def clear_interface(self):
        self.ui.node_name_input.clear()
        self.ui.tac_input.clear()
        self.ui.longitude_input.clear()
        self.ui.node_name_input_5.clear() 
        self.ui.radioButton_700.setChecked(False)
        self.ui.radioButton_700_5m.setChecked(False)
        self.ui.radioButton_1800.setChecked(False)
        self.ui.radioButton_2100.setChecked(False)
        self.ui.radioButton_2600.setChecked(False)
        self.ui.radioButton_2600_tdd.setChecked(False)
        for row in range(self.ui.tableWidget.rowCount()):
            for col in range(self.ui.tableWidget.columnCount()):
                self.ui.tableWidget.setItem(row, col, QtWidgets.QTableWidgetItem(""))

    def save_to_excel(self, export_map):
        template_path = 'template/template.xlsx'
        try:
            filename = f'{export_map.get("Node Name", "output")}_cdr.xlsx'
            shutil.copy(template_path, filename)
            wb = openpyxl.load_workbook(filename)
            ws = wb.active
        except (FileNotFoundError, IOError) as e:
            self.show_message("開啟失敗", f"template 檔開啟失敗: {e}", QMessageBox.Warning)
            return

        for row in range(1, ws.max_row + 1):
            cell_key = ws.cell(row=row, column=1)
            key = str(cell_key.value).strip() if cell_key.value else ""
            if key in export_map:
                value = export_map[key]
                cell_value = ws.cell(
                    row=row,
                    column=2,
                    value=",".join(map(str, value)) if isinstance(value, list) else value
                )
                cell_value.font = Font(color="000000")  

        wb.save(filename)
        self.show_message("匯出成功", "Excel 匯出完成", QMessageBox.Information)


    def is_input_correct(self):
        return True
    
    def show_message(self, title, text, icon):
        msg = QMessageBox(self)
        msg.setWindowTitle(title)
        msg.setText(text)
        msg.setIcon(icon)
        msg.exec_()


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())